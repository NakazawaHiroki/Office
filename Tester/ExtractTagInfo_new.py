from openpyxl import load_workbook
from EdgeStrategy import EdgeStrategy
import re
from selenium.webdriver.common.by import By
import os

SourceFile = "C:/development/Dugong/検査情報_ホクト/ホクト.xlsx"

ExtURL = "http://10.150.80.41:83/dwf-web/login/init"
ExtID = "densan"
ExtPW = "densan"
LedgerID = "90005"
#横幅の最大サイズ
LIMIT_XPOINT = 700.0

g_SelectLists = []

top_value_regex = re.compile(r"top:\s*([^;]+);")
height_value_regex = re.compile(r"height:\s*([^;]+);")
left_value_regex = re.compile(r"left:\s*([^;]+);")

ExtractChoice = [
                    {"tag":"input", "type":"button", "class":"", "proc":"button"},
                    {"tag":"input", "type":"checkbox", "class":"", "proc":"check"},
                    {"tag":"input", "type":"text", "class":"", "proc":"text"},
                    {"tag":"textarea", "type":"", "class":"", "proc":"textarea"},
                    {"tag":"select", "type":"", "class":"", "proc":"select"},
                    {"tag":"table", "type":"", "class":"vi-stamp", "proc":"stamp"},
                ]

# def extractButton(strategy, toppoint, tagHeight, elements):
#     taglist = []
#     top = ''
#     height = ''
#     left = ''
#     for elem in elements:
#         try:
#             style = elem.get_attribute("style")
#             if style and "top:" in style:
#                 top = top_value_regex.search(style).group(1)
#                 height = height_value_regex.search(style).group(1)
#                 left = left_value_regex.search(style).group(1)
#                 if top == toppoint and height == tagHeight and float(left.replace('mm', '')) < LIMIT_XPOINT:
#                     taglist.append([top, height, left, elem.get_attribute("id"), 'button', ''])
#         except:
#             continue
#     return taglist

# #チェックボックス要素を取得す
# def extractCheckButton(strategy, toppoint, tagHeight, elements):
#     taglist = []
#     top = ''
#     height = ''
#     left = ''
#     for elem in elements:
#         try:
#             style = elem.get_attribute("style")
#             if style and "top:" in style:
#                 top = top_value_regex.search(style).group(1)
#                 height = height_value_regex.search(style).group(1)
#                 left = left_value_regex.search(style).group(1)
#                 if top == toppoint and height == tagHeight and float(left.replace('mm', '')) < LIMIT_XPOINT:
#                     taglist.append([top, height, left, elem.get_attribute("id"), 'checkbox', ''])
#         except:
#             continue
#     return taglist

# #text要素を取得する
# def extractText(strategy, toppoint, tagHeight, elements):
#     taglist = []
#     top = ''
#     height = ''
#     left = ''
#     for elem in elements:
#         try:
#             style = elem.get_attribute("style")
#             if style and "top:" in style:
#                 top = top_value_regex.search(style).group(1)
#                 height = height_value_regex.search(style).group(1)
#                 left = left_value_regex.search(style).group(1)
#                 if top == toppoint and height == tagHeight and float(left.replace('mm', '')) < LIMIT_XPOINT:
#                     if elem.get_attribute('readonly') is None:
#                         taglist.append([top, height, left, elem.get_attribute("id"), 'text', ''])
#                     else:
#                         taglist.append([top, height, left, elem.get_attribute("id"), 'readonly', ''])
#         except:
#             continue
#     return taglist

# #text要素を取得する
# def extractTextArea(strategy, toppoint, tagHeight, elements):
#     taglist = []
#     top = ''
#     height = ''
#     left = ''
#     for elem in elements:
#         try:
#             style = elem.get_attribute("style")
#             if style and "top:" in style:
#                 top = top_value_regex.search(style).group(1)
#                 height = height_value_regex.search(style).group(1)
#                 left = left_value_regex.search(style).group(1)
#                 if top == toppoint and height == tagHeight and float(left.replace('mm', '')) < LIMIT_XPOINT:
#                     if elem.get_attribute('readonly') is None:
#                         taglist.append([top, height, left, elem.get_attribute("id"), 'textarea', ''])
#                     else:
#                         taglist.append([top, height, left, elem.get_attribute("id"), 'readonly', ''])
#         except:
#             continue
#     return taglist

# #選択リストを取得する
# def extractSelects(strategy, toppoint, tagHeight, elements):
#     taglist = []
#     top = ''
#     height = ''
#     left = ''
#     for elem in elements:
#         try:
#             style = elem.get_attribute("style")
#             if style and "top:" in style:
#                 top = top_value_regex.search(style).group(1)
#                 height = height_value_regex.search(style).group(1)
#                 left = left_value_regex.search(style).group(1)
#                 if top == toppoint and height == tagHeight and float(left.replace('mm', '')) < LIMIT_XPOINT:
#                     #optionsのリストを抽出する
#                     optionval = []
#                     options = elem.find_elements(By.TAG_NAME, "option")
#                     for option in options:
#                         optionval.append(option.get_attribute("value"))
#                     taglist.append([top, height, left, elem.get_attribute("id"), 'select'])
#                     #選択肢がすでにグローバル変数のリストに保存されているかチェックしてなければ保存する
#                     if not optionval in g_SelectLists:
#                         g_SelectLists.append(optionval)
#         except:
#             continue
#     return taglist



#エクセルファイルをロードする
wb = load_workbook(SourceFile, data_only=True)
sheet = wb.active
initData = []

#ロードしたファイルからリストを作成する A列からI列（1~9列目）
#見出し
for row in sheet.iter_rows(min_row=2, max_col=9, values_only=True): 
    # A, B, C列がすべてNoneであれば終了
    if row[0] is None and row[1] is None and row[2] is None:
        break
    # Noneを空白に変換してリストに追加
    initData.append([cell if cell is not None else '' for cell in row])

#ブラウザを開く
strategy = EdgeStrategy()
strategy.initDriver()
strategy.accessURL(ExtURL)
strategy.login(ExtID, ExtPW)
strategy.openNewLedger(LedgerID)

#ブラウザから抽出する
outputList = []
tagInfoList = []
taglist = []
topPoint = ''
tagHeight = ''
taglistindex = 0
alertFlug = False

all_tags = strategy.enumTag()

for i, row in enumerate(initData, start=1):
    if alertFlug:
        outputList.append(row)  #警告が出ているから以降は元リストのコピー
        continue
    #エクセルのY座標指定が空欄でない、現在ブラウザから抽出しているY座標か高さが不一致ならば次の抽出条件がきた
    if row[3] != '' and (row[3] != topPoint or row[4] != tagHeight):
        #これまでのタグリストとエクセルで予約されているスペースの数が異なるかチェックする
        if len(tagInfoList) > 0 and taglistindex != len(tagInfoList):
            print('エクセルで予約されている行数と異なります')
            alertFlug = True
            outputList.append(row)  #警告が出ているから以降は元リストのコピー
            continue
        if row[3] == '-':
            #'-'が指定されているところは見出しみたいなものなのでコピーだけ
            outputList.append(row)
            continue

        #タグ要素を抽出
        taglistindex = 0
        topPoint = row[3]
        tagHeight = row[4]
        print(f'Y座標:{topPoint}  高さ:{tagHeight} のタグを抽出')

        #タグ選択リストからマッチするタグがあるかチェックする
        for tag in all_tags:
            tag_name = tag.name
            tag_type = tag.get("type", "")  # inputの場合にtype属性を取得
            tag_class = " ".join(tag.get("class", []))  # クラスを文字列に変換
            for choice in ExtractChoice:
                if (choice["tag"] != tag_name or 
                    choice["type"] != tag_type or 
                    (choice["class"] != tag_class and not choice["class"] in tag_class)):
                    continue

                #タグのタイプが一致したので、X,Y座標を確認
                style = tag.get("style", "")
                if style is None or not "top:" in style:
                    continue    #styleが無いtop:が含まれていない
                
                top = top_value_regex.search(style).group(1)
                height = height_value_regex.search(style).group(1)
                left = left_value_regex.search(style).group(1)
                if top == topPoint and height == tagHeight and float(left.replace('mm', '')) < LIMIT_XPOINT:
                    #座標が一致したのでタグのタイプごとに処理を変更する
                    if choice["proc"] == "button":
                        taglist.append([top, height, left, tag.get("id"), 'button', ''])
                    elif choice["proc"] == "check":
                        taglist.append([top, height, left, tag.get("id"), 'checkbox', ''])
                    elif choice["proc"] == "text":
                        if tag.get('readonly') is None:
                            taglist.append([top, height, left, tag.get("id"), 'text', ''])
                        else:
                            taglist.append([top, height, left, tag.get("id"), 'readonly', ''])
                    elif choice["proc"] == "textarea":
                        if tag.get('readonly') is None:
                            taglist.append([top, height, left, tag.get("id"), 'textarea', ''])
                        else:
                            taglist.append([top, height, left, tag.get("id"), 'readonly', ''])
                    elif choice["proc"] == "select":
                        #optionsのリストを抽出する
                        optionval = []
                        options = tag.find_all("option")
                        for option in options:
                            optionval.append(option.get("value"))
                        taglist.append([top, height, left, tag.get("id"), 'select'])
                        #選択肢がすでにグローバル変数のリストに保存されているかチェックしてなければ保存する
                        if not optionval in g_SelectLists:
                            g_SelectLists.append(optionval)
                    elif choice["proc"] == "stamp":
                        #押印はクリックでハンコを付ける、ハンコを消すの二つのタグで動作しているのでどちらか一つをフィルタして追加しない
                        onclick_value = tag.get("onclick")
                        if onclick_value is None or onclick_value.startwith('displayImg'):
                            taglist.append([top, height, left, tag.get("id"), 'stamp'])

        #X座標でソートする
        tagInfoList = sorted(taglist, key=lambda x: float(x[2].replace('mm', '')))

        if len(tagInfoList) == 0:
            print('エクセルで予約されている行数と異なります')
            alertFlug = True
            outputList.append(row)  #警告が出ているから以降は元リストのコピー
            continue
        print('タグの個数は: ' + str(len(tagInfoList)))
    
    #配列サイズとインデックス番号が同数なのは範囲外エラーになるので終了
    if taglistindex == len(tagInfoList):
        print('エクセルで予約されている行数と異なります')
        alertFlug = True
        outputList.append(row)  #警告が出ているから以降は元リストのコピー
    else:
        print(tagInfoList[taglistindex])
        outputList.append(row[:3] + tagInfoList[taglistindex])
        taglistindex += 1

strategy.terminateBrowser()

#ファイルへ保存する 2行目から出力開始
start_row = 2
# outputの内容をシートに書き込む
for i, row in enumerate(outputList, start=start_row):
    for j, value in enumerate(row, start=1):  # 列は1から始まる
        sheet.cell(row=i, column=j, value=value)

# 2番目のシートに<select>タグの選択肢を保存する
selectsheet = wb.create_sheet(title='選択肢のリスト')
for i, options in enumerate(g_SelectLists, start=1):
    for j, option in enumerate(options, start=1):
        selectsheet.cell(row=j, column=i, value=option)

# Excelファイルを別名で保存
file_root, file_ext = os.path.splitext(SourceFile)
# '_Ext'を拡張子の前に追加した新しいファイル名を作成
new_file_path = f"{file_root}_Ext{file_ext}"
# 必要な操作を行う（ここでは既存のデータを書き換えないまま保存）
wb.save(new_file_path)
