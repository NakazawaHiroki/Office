from openpyxl import load_workbook
from EdgeStrategy import EdgeStrategy
import re
from selenium.webdriver.common.by import By
import os

g_SelectLists = []

#指定されたY座標とアイテムの高さのタグ情報を取得する
#X座標で昇順ソートして返却する
def getTagList(strategy, toppoint, tagHeight):
    taglist = []
    top_value_regex = re.compile(r"top: \s*([^;]+);")
    height_value_regex = re.compile(r"height: \s*([^;]+);")
    left_value_regex = re.compile(r"left: \s*([^;]+);")
    top = ''
    height = ''
    left = ''

    #ボタン要素を取得する
    elements = strategy.enumInputTag('button')
    for elem in elements:
        try:
            style = elem.get_attribute("style")
            if style and "top:" in style:
                top = top_value_regex.search(style).group(1)
                height = height_value_regex.search(style).group(1)
                left = left_value_regex.search(style).group(1)
                if top == toppoint and height == tagHeight and float(left.replace('mm', '')) < 550.0:
                    taglist.append([top, height, left, elem.get_attribute("id"), 'button', ''])
        except:
            continue

    #text要素を取得する
    elements = strategy.enumInputTag('text')
    for elem in elements:
        try:
            style = elem.get_attribute("style")
            if style and "top:" in style:
                top = top_value_regex.search(style).group(1)
                height = height_value_regex.search(style).group(1)
                left = left_value_regex.search(style).group(1)
                if top == toppoint and height == tagHeight and float(left.replace('mm', '')) < 550.0:
                    if elem.get_attribute('readonly') is None:
                        taglist.append([top, height, left, elem.get_attribute("id"), 'text', ''])
                    else:
                        taglist.append([top, height, left, elem.get_attribute("id"), 'readonly', ''])
        except:
            continue

    #text要素を取得する
    elements = strategy.enumTextAreaTag()
    for elem in elements:
        try:
            style = elem.get_attribute("style")
            if style and "top:" in style:
                top = top_value_regex.search(style).group(1)
                height = height_value_regex.search(style).group(1)
                left = left_value_regex.search(style).group(1)
                if top == toppoint and height == tagHeight and float(left.replace('mm', '')) < 550.0:
                    if elem.get_attribute('readonly') is None:
                        taglist.append([top, height, left, elem.get_attribute("id"), 'textarea', ''])
                    else:
                        taglist.append([top, height, left, elem.get_attribute("id"), 'readonly', ''])
        except:
            continue

    #選択リストを取得する
    elements = strategy.enumSelectTag()
    for elem in elements:
        try:
            style = elem.get_attribute("style")
            if style and "top:" in style:
                top = top_value_regex.search(style).group(1)
                height = height_value_regex.search(style).group(1)
                left = left_value_regex.search(style).group(1)
                if top == toppoint and height == tagHeight and float(left.replace('mm', '')) < 550.0:
                    #optionsのリストを抽出する
                    optionval = []
                    options = elem.find_elements(By.TAG_NAME, "option")
                    for option in options:
                        optionval.append(option.get_attribute("value"))
                    taglist.append([top, height, left, elem.get_attribute("id"), 'select'])
                    #選択肢がすでにグローバル変数のリストに保存されているかチェックしてなければ保存する
                    if not optionval in g_SelectLists:
                        g_SelectLists.append(optionval)
        except:
            continue

    #X座標でソートする
    result = sorted(taglist, key=lambda x: float(x[2].replace('mm', '')))

    return result



#エクセルファイルをロードする
file_path = "C:/development/Dugong/検査情報_高森町/文書目録起案票.xlsx"
wb = load_workbook(file_path, data_only=True)
sheet = wb.active
initData = []

#ロードしたファイルからリストを作成する A列からI列（1~9列目）
#見出し
for row in sheet.iter_rows(min_row=2, max_col=9, values_only=True): 
    # A, B, C列がすべてNoneであれば終了
    if row[0] is None and row[1] is None and row[2] is None:
        break
    # Noneを空白に変換してリストに追加
    initData.append([cell if cell is not None else '' for cell in row])

#ブラウザを開く
strategy = EdgeStrategy()
strategy.initDriver()
strategy.accessURL("http://172.16.62.86:81/dwf-web/login/init")
strategy.login("densan", "densan")
strategy.openNewLedger("140302")


#ブラウザから抽出する
outputList = []
tagInfoList = []
topPoint = ''
tagHeight = ''
taglistindex = 0
alertFlug = False
for i, row in enumerate(initData, start=1):
    if alertFlug:
        outputList.append(row)  #警告が出ているから以降は元リストのコピー
        continue

    if row[3] != '' and row[3] != topPoint: #Y座標の指定が異なる
        #これまでのタグリストとエクセルで予約されているスペースの数がことなるかチェックする
        if len(tagInfoList) > 0 and taglistindex != len(tagInfoList):
            print('エクセルで予約されている行数と異なります')
            alertFlug = True
            outputList.append(row)  #警告が出ているから以降は元リストのコピー
            continue
        if row[3] == '-':
            #'-'が指定されているところは見出しみたいなものなのでコピーだけ
            outputList.append(row)
            continue
        #タグ要素を抽出
        taglistindex = 0
        topPoint = row[3]
        tagHeight = row[4]
        print(f'Y座標:{topPoint}  高さ:{tagHeight} のタグを抽出')
        tagInfoList = getTagList(strategy, topPoint, tagHeight)
        if len(tagInfoList) == 0:
            print('エクセルで予約されている行数と異なります')
            alertFlug = True
            outputList.append(row)  #警告が出ているから以降は元リストのコピー
            continue
        print('タグの個数は: ' + str(len(tagInfoList)))

    if taglistindex == len(tagInfoList):
        print('エクセルで予約されている行数と異なります')
        alertFlug = True
        outputList.append(row)  #警告が出ているから以降は元リストのコピー
    else:
        print(tagInfoList[taglistindex])
        outputList.append(row[:3] + tagInfoList[taglistindex])
        taglistindex += 1

strategy.terminateBrowser()

#ファイルへ保存する 2行目から出力開始
start_row = 2
# outputの内容をシートに書き込む
for i, row in enumerate(outputList, start=start_row):
    for j, value in enumerate(row, start=1):  # 列は1から始まる
        sheet.cell(row=i, column=j, value=value)

# 2番目のシートに<select>タグの選択肢を保存する
selectsheet = wb.create_sheet(title='選択肢のリスト')
for i, options in enumerate(g_SelectLists, start=1):
    for j, option in enumerate(options, start=1):
        selectsheet.cell(row=j, column=i, value=option)

# Excelファイルを別名で保存
file_root, file_ext = os.path.splitext(file_path)
# '_Ext'を拡張子の前に追加した新しいファイル名を作成
new_file_path = f"{file_root}_Ext{file_ext}"
# 必要な操作を行う（ここでは既存のデータを書き換えないまま保存）
wb.save(new_file_path)
