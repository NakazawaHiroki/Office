import tkinter as tk
from tkinter import Frame, Scrollbar, Button, filedialog, messagebox
from openpyxl import load_workbook

import threading
from enum import IntEnum, auto
from CellCanvas import CellCanvas
from EdgeStrategy import EdgeStrategy
import CONST
import os

###################################################################
#   アプリケーション定数
###################################################################
WINDOW_WIDTH            = 1200
WINDOW_HEIGHT           = 700
BUTTON_CAPTION_SIZE     = 9
BUTTON_CAPTION_FONT     = 'Arial'
LABEL_TEXT_SIZE         = 9
LABEL_TEXT_FONT         = 'Arial'
#########################################
#テストデータに関する設定
EXCEL_LOAD_SHEET_NAME   = 'Sheet1'      #ロードするエクセルファイルで優先的にロードするシート名
TAGID_COLUMN            = 6             #Webページ上のボタンやテキストなどのIDが入力された列の位置
TAGTYPE_COLUMN          = 7             #Webページ上のボタンやテキスト、リストなどの種類が入力された列の位置
NOLOAD_COLUMN           = [3,4,5,6,7]   #リスト上に表示しないエクセルファイルの列番号(タグIDや種類など)
class LIST_FIX(IntEnum):                #テストデータではない固定の行(左上リストに表示する内容)
    LOGIN           = 1
    LEDGERID        = auto()
    END             = auto()
    ERROR           = auto()
#########################################
#リストに関する設定
LIST_TOP_POSITION       = 85            #リストの表示位置


###################################################################
#   AutoInput
###################################################################
class AutoInput:
    def __init__(self, root, strategy):
        self.m_root: tk.Tk                  = root
        self.browStrategy: EdgeStrategy     = strategy
        self.listFrame4: Frame              = None
        self.listFrame5: Frame              = None
        self.leftList: CellCanvas           = None
        self.rightList: CellCanvas          = None
        self.topLeftList: CellCanvas        = None
        self.topRightList: CellCanvas       = None
        self.vScroll                        = None
        self.hScroll                        = None
        self.tagTypeList                    = []
        self.tagIDList                      = []
        self.flagRun                        = False
        self.runThread                      = None
        self.colCount                       = 0
        self.lockWidget                     = []
        self.versionLabel                   = None

        # メインフレームを作成
        main_frame = Frame(root)
        main_frame.pack(fill=tk.BOTH, expand=True)
        #ボタン等の作成
        self.createComponent(main_frame)
        #サブフレームの作成
        self.createFrame(main_frame)

        #初期設定をロードする
        self.loadRegistry()

        #マウスのホイールボタンイベント
        main_frame.bind_all("<MouseWheel>", self.on_MouseWheelRotate)
        main_frame.bind_all("<Shift-MouseWheel>", self.on_MouseWheelRotateShift)

        #ファイルのロードとリストの作成
        # self.reloadTestData()


    def load_excel_data(self, filepath, excluded_columns=None):
        # Excelファイルを読み込む
        workbook = load_workbook(filepath, data_only=True)
        if EXCEL_LOAD_SHEET_NAME in workbook.sheetnames:
            sheet = workbook[EXCEL_LOAD_SHEET_NAME]
        else:
            sheet = workbook.active

        leftdata        = []
        rightdata       = []
        topLeftData     = []
        topRightData    = []
        tagTypes        = []
        tagIDs          = []
        excluded_columns = excluded_columns or []  # デフォルトで除外列がない場合は空リスト
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            # 指定された列を除外
            filtered_row = [value if value is not None else "" for idx, value in enumerate(row) if idx not in excluded_columns]
            #配列は0から始まっているので、上下の切り分けでは1加算する
            if i <= (list(LIST_FIX)[-1].value + 1):
                topLeftData.append([i] + filtered_row[:3])
                topRightData.append(filtered_row[3:])
            else:
                leftdata.append([i] + filtered_row[:3])
                rightdata.append(filtered_row[3:])
                # タグ属性とタグIDを取得する
                tagIDs.append(row[TAGID_COLUMN])
                tagTypes.append(row[TAGTYPE_COLUMN])
                
        return leftdata, rightdata, topLeftData, topRightData, tagTypes, tagIDs


    def reloadTestData(self):
        #パスを取得
        loadpath = self.filepath_entry.get()

        if loadpath == '':
            return

        #リストを破棄する
        if self.topLeftList is not None:
            self.topLeftList.destroy()
            self.topLeftList = None
        #リストを破棄する
        if self.topRightList is not None:
            self.topRightList.destroy()
            self.topRightList = None
        #リストを破棄する
        if self.leftList is not None:
            self.leftList.destroy()
            self.leftList = None
        #リストを破棄する
        if self.rightList is not None:
            self.rightList.destroy()
            self.rightList = None

        #ファイルの有無をチェックする
        if not os.path.isfile(loadpath):
            return
        # エクセルファイルを読み込む
        filepath = loadpath
        leftdata, rightdata, topleftdata, toprightdata, tagTypes, tagIDs = self.load_excel_data(filepath, NOLOAD_COLUMN)

        self.tagTypeList = tagTypes
        self.tagIDList = tagIDs

        #リストの作成
        leftColCount = len(leftdata[0])
        rightColCount = len(rightdata[0])
        self.colCount = len(rightdata[0])

        cell_height = 18
        # 列ごとの幅を指定
        left_column_widths = [30, 85, 85, 85]
        right_column_widths = [75] * rightColCount

        #左上のリストを作成する(水平スクロール固定)
        rows = len(topleftdata)
        self.topLeftList = CellCanvas(self.listFrame4, False, False, rows, leftColCount, left_column_widths, cell_height, 
                            width=sum(left_column_widths), height=rows*cell_height,
                            scrollregion=(0, 0, sum(left_column_widths), rows*cell_height))
        self.topLeftList.pack(side=tk.TOP, fill=tk.X)

        #左のリストを作成する(水平スクロール固定)
        rows = len(leftdata)
        self.leftList = CellCanvas(self.listFrame4, False, False, rows, leftColCount, left_column_widths, cell_height, 
                            width=sum(left_column_widths), height=rows*cell_height,
                            scrollregion=(0, 0, sum(left_column_widths), (rows)*cell_height))
        self.leftList.pack(side=tk.BOTTOM, fill=tk.X)

        #右上のリストを作成する
        rows = len(toprightdata)
        self.topRightList = CellCanvas(self.listFrame5, True, False, rows, rightColCount, right_column_widths, cell_height, 
                            width=sum(right_column_widths), height=rows*cell_height,
                            scrollregion=(0, 0, sum(right_column_widths), (rows)*cell_height))
        self.topRightList.pack(side=tk.TOP, fill=tk.X)

        #右のリストを作成する
        rows = len(rightdata)
        self.rightList = CellCanvas(self.listFrame5, True, True, rows, rightColCount, right_column_widths, cell_height, 
                            width=sum(right_column_widths), height=rows*cell_height,
                            scrollregion=(0, 0, sum(right_column_widths), (rows)*cell_height))
        self.rightList.pack(side=tk.BOTTOM, fill=tk.X)
        self.rightList.setNotifyVScroll(self.leftList)

        #スクロールイベントを設定
        self.rightList.configure(xscrollcommand=self.hScroll.set)
        self.topRightList.configure(xscrollcommand=self.hScroll.set)
        self.leftList.configure(yscrollcommand=self.vScroll.set)
        self.rightList.configure(yscrollcommand=self.vScroll.set)

        #クリックイベントを設定
        self.topRightList.setNotifyClick(self.rightList.fireClickEvent)
        self.rightList.setNotifyClick(self.topRightList.fireClickEvent)

        # エクセルデータをCanvasにロード
        self.topLeftList.load_data(topleftdata)
        self.leftList.load_data(leftdata)
        self.topRightList.load_data(toprightdata)
        self.rightList.load_data(rightdata)

        #レジストリに今の情報を保存
        self.saveRegistry()


    def open_file(self):
        initialdir = ''
        loadpath = self.filepath_entry.get()
        if loadpath != '':
            if os.path.exists(os.path.dirname(loadpath)) and os.path.isdir(os.path.dirname(loadpath)):
                initialdir = os.path.dirname(loadpath)

        if initialdir == '':
            initialdir = os.getcwd()

        filetypes = (
            ("Excel files", "*.xlsx"),
            ("All files", "*.*")
        )
        filepath = filedialog.askopenfilename(
            title="テスト用ファイルを選んでください",
            initialdir=initialdir,
            filetypes=filetypes
        )
        if filepath:
            self.filepath_entry.delete(0, tk.END)  # テキストボックスをクリア
            self.filepath_entry.insert(0, filepath)  # ファイルパスを表示
            self.reloadTestData()


    def _on_StepRun(self):
        if self.rightList is None or self.rightList.get_select_col() is None:
            return
        self.flagRun = True
        self.enterRun(True)
        self.runThread = threading.Thread(target=self._currentColRun)
        self.runThread.start()

    def _on_Run(self):
        if self.rightList is None or self.rightList.get_select_col() is None:
            return
        self.flagRun = True
        self.enterRun(True)
        self.runThread = threading.Thread(target=self._batchProcessing)
        self.runThread.start()

    def _on_TerminateRun(self):
        self.flagRun = False
        if hasattr(self, 'thread'):
            self.runThread.join()

    #現在の選択列から最後まで実行する
    def _batchProcessing(self):
        try:
            if self.rightList is None or self.rightList.get_select_col() is None:
                return
            col = self.rightList.get_select_col()
            if col is None:
                return
            
            for i in range(col, self.colCount):
                if self.check_followCell.get() == 1:
                    self.topRightList.select_col(i, True)
                    self.rightList.select_col(i, True)
                else:
                    self.topRightList.select_col(i, False)
                    self.rightList.select_col(i, False)
                self._currentColRun(True)
                if not self.flagRun:
                    break
        finally:
            self.enterRun(False)


    #現在選択中の列を実行する
    def _currentColRun(self, batchproc = False):
        try:
            #先頭を表示する
            #入力中セルを追従する場合はセルを表示するまでスクロール
            if self.check_followCell.get() == 1:
                self.rightList.display_row(0)
                self.leftList.display_row(0)
            # 現在選択されている列の値を取得して出力
            topselected_values = self.topRightList.get_selected_column_values()
            selected_values = self.rightList.get_selected_column_values()
            #最初の番号以外は全て空白なら終了
            if all(item == '' for item in topselected_values[1:]):
                if all(item == '' for item in selected_values):
                    self.flagRun = False
                    return

            #ログインユーザーを確認して現在のユーザーと異なる指定があればそっちでログインする
            #リストで指定されたログインIDとパスワードと帳票ID
            login = topselected_values[LIST_FIX.LOGIN]
            openLedgerID = topselected_values[LIST_FIX.LEDGERID]

            if openLedgerID == '' and not self.browStrategy.IsBrowserOpen():
                self.flagRun = False
                messagebox.showwarning('警告', '帳票を開くボタンを押して帳票を開いてください')
                return

            #アプリ上で指定されたログインIDとパスワードと帳票ID
            specifyID = self.ID_entry.get()
            specifyPass = self.pass_entry.get()
            specifyLedgerID = self.ledger_entry.get()

            if login != '' and ',' in login:
                tempID = login.split(',')[0]
                tempPASS = login.split(',')[1]
                if '=' in tempID and '=' in tempPASS:
                    specifyID = tempID.split('=')[1]
                    specifyPass = tempPASS.split('=')[1]

            if openLedgerID != '':
                specifyLedgerID = openLedgerID

            if (self.browStrategy.getLogiID() != specifyID
                or self.browStrategy.getLoginPass() != specifyPass
                or self.browStrategy.getOpenLedger() != specifyLedgerID):
                result = self._OpenLedger(self.url_entry.get(),
                                specifyID,
                                specifyPass,
                                False,
                                specifyLedgerID)
                if result != CONST.RES_SUCCESS:
                    self.flagRun = False
                    return

            #データを入力する
            tagType = ""
            tagID = ""
            for i, value in enumerate(selected_values):
                if not self.flagRun:
                    break   #処理の停止

                #タグ属性とタグIDを取得する
                tagType = self.tagTypeList[i]
                tagID = self.tagIDList[i]
                #タグIDへ値をセットする
                if tagID != '' and tagID != '-' and tagType != 'readonly' and value != '':
                    #入力中セルを追従する場合はセルを表示するまでスクロール
                    if self.check_followCell.get() == 1:
                        self.rightList.display_row(i)
                        self.leftList.display_row(i)
                    result = self.browStrategy.setValue(tagType, tagID, value)
                    self.rightList.reflectionCell(result, i, self.rightList.get_select_col())

            #入力後の動作を確認して実行
            if self.flagRun:
                if topselected_values[LIST_FIX.END] == '入力チェック':
                    result = self.browStrategy.fireInputCheck()
                    if result != CONST.RES_SUCCESS:
                        #エラー内容を取得して表示する
                        errorCode, errorMessage = self.browStrategy.getErrorInfo()
                        mes = f'{errorCode} : {errorMessage}'
                        self.topRightList.updateCellText(mes, LIST_FIX.ERROR, self.topRightList.get_select_col(), True)
                        self.browStrategy.fireGoBack()
                elif topselected_values[LIST_FIX.END] == '起票':
                    result = self.browStrategy.fireStartReport()
                    if result == CONST.RES_BUTTON_NONE:
                        messagebox.showwarning('警告', '起票ボタンはありません')
                        self.flagRun = False
                    return
                elif topselected_values[LIST_FIX.END] == '承認':
                    result = self.browStrategy.fireAcceptReport()
                    if result == CONST.RES_BUTTON_NONE:
                        messagebox.showwarning('警告', '承認ボタンはありません')
                        self.flagRun = False
                    elif result == CONST.RES_ERROR:
                        errorCode, errorMessage = self.browStrategy.getErrorInfo()
                        mes = f'{errorCode} : {errorMessage}'
                        self.topRightList.updateCellText(mes, LIST_FIX.ERROR, self.topRightList.get_select_col(), True)
                        self.browStrategy.fireGoBack()
                        self.flagRun = False
                    return
                else:
                    return

            #readOnlyの属性情報を比較する
            for i, value in enumerate(selected_values):
                if self.flagRun == False:
                    break   #処理の停止
                if self.check_emptyCell.get() == 0 and value == '':
                    continue

                #タグ属性とタグIDを取得する
                tagType = self.tagTypeList[i]
                tagID = self.tagIDList[i]
                #タグIDへ値をセットする
                if tagID != '' and tagID != '-' and tagType == 'readonly':
                    self.rightList.display_row(i)
                    self.leftList.display_row(i)
                    result = self.browStrategy.compareValue(tagID, value)
                    self.rightList.reflectionCell(result, i, self.rightList.get_select_col())
        finally:
            if not batchproc:
                self.enterRun(False)

    #実行の切り替え指定
    def enterRun(self, enterProc):
        self.rightList.enterProcess(enterProc)

        wiState = tk.DISABLED
        if not enterProc:
            wiState = tk.NORMAL
        for elem in self.lockWidget:
            elem.config(state=wiState)


    #帳票を開くボタンが押された
    def _on_OpenLedger(self):
        url = self.url_entry.get()
        id = self.ID_entry.get()
        pw = self.pass_entry.get()
        bNew = False
        if self.OpenLedgerType.get() == 1:
            bNew = True
        ledgerID = self.ledger_entry.get()
        self.flagRun = True
        self._OpenLedger(url, id, pw, bNew, ledgerID)
        self.flagRun = False


    #帳票を開くボタンが押された
    def _OpenLedger(self, url, id, pw, bNew, ledgerID):
        self.m_root.config(cursor='watch')
        try:
            self.m_root.update()
            self.browStrategy.terminateBrowser()
            if not self.flagRun: return CONST.RES_STOP
            result = self.browStrategy.initDriver()
            if result == CONST.RES_WEBDRV_NONE:
                messagebox.showwarning('警告', 'ドライバが見つかりません\nmsedgedriver.exeを実行ファイルと同じ場所に保存してください')
                return result
            elif result == CONST.RES_WEBDRV_INCOMPATIBLE:
                messagebox.showwarning('警告', 'ドライバとブラウザのバージョンが一致しません')
                return result
            elif result == CONST.RES_WEBDRV_LOADERROR:
                messagebox.showwarning('警告', 'ドライバのロードに失敗しました')
                return result
            if not self.flagRun: return CONST.RES_STOP
            result = self.browStrategy.accessURL(url)
            if not self.flagRun: return CONST.RES_STOP

            if result == CONST.RES_SUCCESS:
                result = self.browStrategy.login(id, pw)
                if result == CONST.RES_SUCCESS:
                    if not self.flagRun:
                        return CONST.RES_STOP
                    if bNew:
                        result = self.browStrategy.openNewLedger(ledgerID)
                    else:
                        result = self.browStrategy.openUnProcessLedger(ledgerID)
        finally:
            self.m_root.config(cursor='')

        if result == CONST.RES_NO_LEDGER:
            messagebox.showwarning('警告', '指定したIDの帳票はありません')
        elif result == CONST.RES_SUCCESS:
            self.saveRegistry()
        return result


    #ウィンドウ上でマウスのホイールが回転した
    def on_MouseWheelRotate(self, event):
        if self.leftList is not None and self.rightList is not None:
            self.leftList.on_mousewheel(event)
            self.rightList.on_mousewheel(event)

    #ウィンドウ上でShiftボタンが押されながらマウスのホイールが回転した
    def on_MouseWheelRotateShift(self, event):
        if self.rightList is not None and self.topRightList is not None:
            self.rightList.on_shift_mousewheel(event)
            self.topRightList.on_shift_mousewheel(event)

    #ウィンドウ内のコンポネントを作成する
    def createComponent(self, main_frame):
        #帳票を開く情報のコンポネント
        # URL表示用のテキストボックス
        label = tk.Label(main_frame, text='URL:', font=(LABEL_TEXT_FONT, LABEL_TEXT_SIZE))
        label.place(x=4, y=0)
        self.url_entry = tk.Entry(main_frame, width=50, font=('Arial', 10))
        self.url_entry.place(x=40, y=2)

        # ID表示用のテキストボックス
        label = tk.Label(main_frame, text='ID:', font=(LABEL_TEXT_FONT, LABEL_TEXT_SIZE))
        label.place(x=16, y=25)
        self.ID_entry = tk.Entry(main_frame, width=15, font=('Arial', 10))
        self.ID_entry.place(x=40, y=27)

        # パスワード表示用のテキストボックス
        label = tk.Label(main_frame, text='PW:', font=(LABEL_TEXT_FONT, LABEL_TEXT_SIZE))
        label.place(x=8, y=49)
        self.pass_entry = tk.Entry(main_frame, width=15, font=('Arial', 10))
        self.pass_entry.place(x=40, y=51)

        # 帳票フォームID表示用のテキストボックス
        label = tk.Label(main_frame, text='帳票ID:', font=(LABEL_TEXT_FONT, LABEL_TEXT_SIZE))
        label.place(x=170, y=50)
        self.ledger_entry = tk.Entry(main_frame, width=13, font=('Arial', 10))
        self.ledger_entry.place(x=215, y=52)

        #新規か未処理帳票か選択するラジオボタン
        self.OpenLedgerType = tk.IntVar()
        radio1 = tk.Radiobutton(root, text="新規帳票", variable=self.OpenLedgerType, value=1, font=(BUTTON_CAPTION_FONT, BUTTON_CAPTION_SIZE))
        radio2 = tk.Radiobutton(root, text="未処理帳票", variable=self.OpenLedgerType, value=2, font=(BUTTON_CAPTION_FONT, BUTTON_CAPTION_SIZE))
        radio1.place(x=165, y=24)
        radio2.place(x=245, y=24)

        openLedgerButton = Button(main_frame, text="帳票を開く", 
                               command=lambda: self._on_OpenLedger(), 
                               bg="RoyalBlue1", font=(BUTTON_CAPTION_FONT, BUTTON_CAPTION_SIZE))
        openLedgerButton.place(x=320, y=50)
        self.lockWidget.append(openLedgerButton)


        # 空欄のセルは比較をスキップするチェックボックスを作成する
        self.check_emptyCell = tk.IntVar()
        check_emptyCell_btn = tk.Checkbutton(root, text="空セルも比較する", variable=self.check_emptyCell, font=(BUTTON_CAPTION_FONT, BUTTON_CAPTION_SIZE))
        check_emptyCell_btn.place(x=420, y=35)

        # 実行中のセルを追従するチェックボックス
        self.check_followCell = tk.IntVar()
        self.check_followCell.set(1)
        check_followCell_btn = tk.Checkbutton(root, text="実行中セルを追従", variable=self.check_followCell, font=(BUTTON_CAPTION_FONT, BUTTON_CAPTION_SIZE))
        check_followCell_btn.place(x=420, y=57)

        stepRunButton = Button(main_frame, text="1列だけ実行", 
                               command=lambda: self._on_StepRun(), 
                               bg="cyan", font=(BUTTON_CAPTION_FONT, BUTTON_CAPTION_SIZE))
        stepRunButton.place(x=560, y=45)

        RunButton = Button(main_frame, text="カーソル列から実行", 
                           command=lambda: self._on_Run(), 
                           bg="cyan", font=(BUTTON_CAPTION_FONT, BUTTON_CAPTION_SIZE))
        RunButton.place(x=660, y=45)
        self.lockWidget.append(stepRunButton)
        self.lockWidget.append(RunButton)

        terminateButton = Button(main_frame, text="停  止", 
                                 command=lambda: self._on_TerminateRun(), 
                                 bg="red", font=(BUTTON_CAPTION_FONT, BUTTON_CAPTION_SIZE))
        terminateButton.place(x=800, y=45)

        # ファイルを開くボタン
        self.open_button = Button(main_frame,  text="開く...", 
                                 command=lambda: self.open_file(), 
                               font=(BUTTON_CAPTION_FONT, BUTTON_CAPTION_SIZE))
        self.open_button.place(x=880, y=0)
        self.lockWidget.append(self.open_button)

        # ファイルパス表示用のテキストボックス
        self.filepath_entry = tk.Entry(main_frame, width=75)
        self.filepath_entry.place(x=420, y=10)

        reloadButton = Button(main_frame, text="読込み", 
                                 command=lambda: self.reloadTestData(), 
                                 bg="yellow", font=(BUTTON_CAPTION_FONT, BUTTON_CAPTION_SIZE))
        reloadButton.place(x=880, y=25)
        self.lockWidget.append(reloadButton)

        #ドライバー、ブラウザーのバージョンを表示する
        bver = self.browStrategy.getBrowserVer()
        dver = self.browStrategy.getDriverVer()
        self.versionLabel = tk.Label(main_frame, text=f'ブラウザVer:{bver}\nドライバVer:{dver}', width=40, font=(LABEL_TEXT_FONT, LABEL_TEXT_SIZE), anchor="w")
        self.versionLabel.place(x=940, y=0)

    #サブフレームとスクロールバーの作成
    def createFrame(self, main_frame):
        listFrame1 = Frame(main_frame, bg="gray", )
        listFrame1.place(x=0, 
                            y=LIST_TOP_POSITION, 
                            relwidth=1, 
                            relheight=(WINDOW_HEIGHT - LIST_TOP_POSITION) / WINDOW_HEIGHT)

        # 垂直スクロールバーの同期
        def sync_vscroll(*args):
            if self.leftList is not None:
                self.leftList.yview(*args)
                self.leftList.adjust_allow()
            if self.rightList is not None:
                self.rightList.yview(*args)
                self.rightList.adjust_allow()

        # 水平スクロールバーの同期
        def sync_hscroll(*args):
            if self.topRightList is not None:
                self.topRightList.xview(*args)
                self.topRightList.adjust_allow()
            if self.rightList is not None:
                self.rightList.xview(*args)
                self.rightList.adjust_allow()

        # 水平スクロールバーを作成
        self.hScroll = Scrollbar(listFrame1, orient=tk.HORIZONTAL, command=sync_hscroll)
        self.hScroll.pack(side=tk.BOTTOM, fill=tk.X)

        #サブフレームの作成
        listFrame2 = Frame(listFrame1, bg="gray")
        listFrame2.pack(side=tk.TOP, fill=tk.X)

        # 垂直スクロールバーを作成
        self.vScroll = Scrollbar(listFrame2, orient=tk.VERTICAL, command=sync_vscroll)
        self.vScroll.pack(side=tk.RIGHT, fill=tk.Y)

        #サブフレームの作成
        listFrame3 = Frame(listFrame2, bg="gray")
        listFrame3.pack(side=tk.LEFT, fill=tk.Y)

        #リストを格納するフレームの作成
        self.listFrame4 = Frame(listFrame3, bg="gray")
        self.listFrame4.pack(side=tk.LEFT, fill=tk.Y)
        self.listFrame5 = Frame(listFrame3, bg="gray")
        self.listFrame5.pack(side=tk.RIGHT, fill=tk.Y)


    def saveRegistry(self):
        # 実行ディレクトリに `config.inf` ファイルを作成
        file_path = os.path.join(os.getcwd(), "config.inf")
        config_data = {}
        try:
            with open(file_path, 'r') as config_file:
                for line in config_file:
                    if '=' in line:
                        key, value = line.strip().split('=', 1)
                        config_data[key] = value

            # "TestURL"の値を更新
            config_data['TestURL']          = self.url_entry.get()
            config_data['TestID']           = self.ID_entry.get()
            config_data['TestPassword']     = self.pass_entry.get()
            config_data['LedgerID']         = self.ledger_entry.get()
            config_data['LedgerType']       = self.OpenLedgerType.get()
            config_data['TestFilepath']     = self.filepath_entry.get()

            # ファイルに書き込み
            with open(file_path, 'w') as config_file:
                for key, value in config_data.items():
                    config_file.write(f"{key}={value}\n")
        
        except Exception as e:
            print(f"エラーが発生しました: {e}")


    def loadRegistry(self):
        file_path = os.path.join(os.getcwd(), "config.inf")
        try:
            with open(file_path, 'r') as config_file:
                for line in config_file:
                    key, value = line.strip().split('=')
                    if key == "TestURL":
                        self.url_entry.insert(0, value)
                    elif key == "TestID":
                        self.ID_entry.insert(0, value)
                    elif key == "TestPassword":
                        self.pass_entry.insert(0, value)
                    elif key == "LedgerID":
                        self.ledger_entry.insert(0, value)
                    elif key == "LedgerType":
                        self.OpenLedgerType.set(int(value))
                    elif key == "TestFilepath":
                        self.filepath_entry.insert(0, value)
        except FileNotFoundError:
            self.url_entry.insert(0, "http://10.150.16.158:84/dwf-web/login/init")
            self.ID_entry.insert(0, "densan")
            self.pass_entry.insert(0, "densan")
            self.ledger_entry.insert(0, "000000")
            self.OpenLedgerType.set(2)


#メインループ
if __name__ == "__main__":
    root = tk.Tk()
    root.title("自動テストツール")
    root.geometry(f'{WINDOW_WIDTH}x{WINDOW_HEIGHT}')  # ウィンドウの初期サイズを設定

    #ブラウザを起動する
    strategy = EdgeStrategy()
    #アプリを起動する
    app = AutoInput(root, strategy)
    root.update_idletasks()
    root.mainloop()
    #最後にブラウザのプロセスの終了処理
    strategy.terminateBrowser()
